import io
import urllib
import zipfile
import datetime

import openpyxl
from openpyxl.writer.excel import save_virtual_workbook

from .vendor_fetch import vendors
from .utils import chunks, sanitize_filename


class UNLRequisition:
    _blank_form_raw = None

    def __init__(self, order):
        self.order = order

    def fetch_empty_form(self):
        """
        Fetches an blank requisition form from the physics website
        """
        global _blank_form_raw
        blank_req_url = "http://www.unl.edu/physics/docs/Requisition2014.xlsx"
        if self._blank_form_raw is None:
            self._blank_form_raw = urllib.request.urlopen(blank_req_url).read()
        bio = io.BytesIO(self._blank_form_raw)
        self.workbook = openpyxl.load_workbook(bio)

    def populate_misc_fields(self, vendor):
        ws = self.workbook.active
        ws['B11'] = vendor.form_info['long_name']
        ws['B13'] = vendor.form_info['addr_line1']
        ws['B16'] = vendor.form_info['addr_line2']
        ws['B18'] = self.order.description
        ws['B22'] = vendor.form_info['phone']
        ws['F22'] = vendor.form_info['fax']
        ws['B26'] = vendor.form_info['website']
        ws['E28'] = str(self.order.delivery_date)
        ws['D43'] = self.order.cost_object
        ws['B45'] = datetime.date.today().strftime("%b. %d, %Y")
        ws['C47'] = self.order.requestor_name
        ws['K47'] = self.order.requestor_phone
        ws['C49'] = self.order.supervisor_name

    def populate_parts(self, parts):
        ws = self.workbook.active
        for i, part in enumerate(parts):
            i = str(i+32)
            ws["A"+i] = part.vendorpart.vendor_part_number
            ws["B"+i] = part.vendorpart.part.short_description
            n = part.number_ordered
            ws["L"+i] = n
            try:
                ws["Q"+i] = part.vendorpart.get_pricebreak(n)
            except ValueError:
                ws["Q"+i] = "N/A"

    def place_sheet_number(self, i, n):
        ws = self.workbook.active
        ws["A52"] = "Sheet {} of {}".format(i, n)

    def get_form(self):
        """
        Creates a zip file containing the one or more xlsx formatted
        requisitions as required by UNL Physics purchasing. Creates the file
        in memory and then returns it as bytes.
        """
        parts_chunked = chunks(self.order.vendorparts, 10)
        bio = io.BytesIO()
        zf = zipfile.ZipFile(bio, 'x')
        for i, parts_chunk in enumerate(parts_chunked):
            self.fetch_empty_form()
            self.populate_parts(parts_chunk)
            self.populate_misc_fields(vendors['Digikey'])
            self.place_sheet_number(i+1, len(parts_chunked))

            fname = "digikey_req{:02d}.xlsx".format(i+1)
            zf.writestr(fname, save_virtual_workbook(self.workbook))
        zf.close()
        bio.seek(0)
        fname = sanitize_filename(self.order.order_name, "zip")
        return fname, bio.read()


class DigikeyCart:

    def __init__(self, order):
        self.order = order

    def get_form(self):
        lines = ["{},{},".format(vendorpart.vendorpart.vendor_part_number,
                                 vendorpart.number_ordered)
                 for vendorpart in self.order.vendorparts
                 if vendorpart.vendorpart.vendor == "Digikey"]
        cart = "\n".join(lines)
        fname = sanitize_filename(self.order.order_name, "csv")
        return fname, cart.encode('utf8')

if __name__ == '__main__':
    from .models import Order
    order = Order.query.all()[-1]
    with open('outfile.zip', 'wb') as f:
        req = UNLRequisition(order)
        f.write(req.get_form(order))
